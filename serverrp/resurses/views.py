from .models import Otdel , Zaiyvka , Contacts , Zadatha ,Cotrudniri , PriseofmyWok ,OtchetforMonf ,Compani
from projects.models import Projects , Itap
from preseil.models import PresailProjects
from django.contrib.auth.models import User

from django.contrib.auth.decorators import login_required
from django.conf.global_settings import MEDIA_ROOT
from django.shortcuts import render, redirect, get_object_or_404, get_list_or_404
from django.core.mail import send_mail , send_mass_mail
from django.db.models import Q
from django.http import FileResponse

from openpyxl import Workbook  , load_workbook
from openpyxl.writer.excel import save_workbook

from datetime import datetime

import os

# Create your views here.
# УДАЛИТЬ _ ПЕРЕДЕЛАТЬ
def main(request):
    pass

# УДАЛИТЬ _ ПЕРЕДЕЛАТЬ
def contacts(request):
    otdel = Otdel.objects.all()
    cont = Contacts.objects.all()
    return render(request, 'resurses/prodject-main.html',{"cont":cont, "otdel":otdel})

# УДАЛИТЬ _ ПЕРЕДЕЛАТЬ
def mycontacts(request, pk):
    #prog = get_object_or_404(Projects, id = pk)
    #otdel = Otdel.objects.filter(projects__title=prog)
    cont = Contacts.objects.all()

    return render(request, 'resurses/prodject-mycontacts.html',{"cont":cont})

#
@login_required(login_url='/lg/')
def mycontactcreate(request, pk):
    prog = get_object_or_404(Projects, id = pk)
    otdel = Otdel.objects.filter(projects__title=prog)
    cont = Contacts.objects.all()

    if request.method == 'POST':
        otdelid = request.POST.getlist('otdelid')
        otdellevel = request.POST.getlist('otdellevel')
        otdeltitle = request.POST.getlist('otdeltitle')
        otdelopisanie = request.POST.getlist('otdelopisanie')
       
        f=0 
        for i in otdelid:
            ot = Otdel.objects.get(id = i)
            ot.title = otdeltitle[f]
            ot.level = otdellevel[f]
            ot.opisanie = otdelopisanie[f]
            ot.save()
            f = f+ 1

        contid = request.POST.getlist('otdelid')
        conttelnomder = request.POST.getlist('conttelnomder')
        conttitle = request.POST.getlist('conttitle')
        contmail = request.POST.getlist('contmail')
        contdoljnost = request.POST.getlist('contdoljnost')
        try:
            g = 0
            for u in contid:
                cont = Contacts.objects.get(id = u)
                cont.telnomder = conttelnomder[g]
                cont.title = conttitle[g]
                cont.mail = contmail[g]
                cont.doljnost = contdoljnost[g]
                cont.save()
                g =g +1 
        except :
            pass
        
        fio = request.POST.getlist('fio')
        telnomnew = request.POST.getlist('telnomnew')
        mailnew = request.POST.getlist('mailnew')
        doljn = request.POST.getlist('doljn')
        iddd = request.POST.getlist('iddd')
        try:
            g = 0
            for inr in iddd:
                cont = Contacts()
                cont.telnomder = telnomnew[g]
                cont.title = fio[g]
                cont.mail = mailnew[g]
                cont.doljnost = doljn[g]
                cont.ondel = Otdel.objects.get(id = inr)
                cont.save()
                g = g+ 1
            
        except:
            pass

        return redirect('projects:views')

    return render(request, 'resurses/prodject-mycontacts-rename.html',{"cont":cont, "otdel":otdel})

# МОИ ЗАДАЧИ 
@login_required(login_url='/lg/')
def myplans(request):
    proj = Projects.objects.filter(rpproecta=request.user.first_name)
    zada = Zadatha.objects.filter(Q(user__first_name=request.user.first_name)|Q(projects__in = proj))
    projd = Projects.objects.all()
    use  = User.objects.all()

    if request.method == 'POST':
        tupe = request.POST.get('tupe')
        if tupe =="new":
            #text = request.POST.get('text')
            le = request.POST.get('projid')
            za = Zadatha()
            za.title = request.POST.get('title')
            za.zada = request.POST.get('text')
            if le == "---":
                pass
            else:
                prod = Projects.objects.get(id=le)
                za.projects = prod
            za.user = request.user
            za.prior = request.POST.get('prior')
            za.srok = request.POST.get('srok')
            za.otvet = request.POST.get('otvet')
            za.save()
        elif tupe == 'old':
            za = Zadatha.objects.get(id = request.POST.get('zadid'))
            za.title = request.POST.get('title')
            za.zada = request.POST.get('text')
            za.prior = request.POST.get('prior')
            za.srok = request.POST.get('srok')
            za.otvet = request.POST.get('otvet')
            za.save()
        elif tupe == 'chec':
            checch = request.POST.getlist('checch')
            for i in checch:
                za = Zadatha.objects.get(id = i)
                za.chec = 'checked'
                za.save()
        elif tupe == 'newtasks':
            prod = Projects.objects.get(id=request.POST.get('projid'))
            #s = request.POST.get('otvet') 
            if  request.POST.get('otvet') == '---':
                us = User.objects.get(first_name=prod.rpproecta)
            else:
                us = User.objects.get(first_name=request.POST.get('otvet'))

            za = Zadatha()
            za.title = request.POST.get('title')
            text = request.POST.get('text')
            za.zada = text
            za.projects = prod
            za.user = us
            za.prior = request.POST.get('prior')
            za.srok = request.POST.get('srok')

            za.otvet = request.POST.get('otvet')
            za.save()
            
            # send_mail(
            #     f'[Системы РМО] Задача по проекту {prod.title}',
            #     f'Новая задача по Проекту:{prod.title}.\r\n  \r\n {text} \r\n  \r\n Системы РМО https://ru01-app38.galaktika.ru/',
            #     'system.pmo@galaktika.ru',
            #     [us.email],
            #     fail_silently=False, 
            #     )
      
        else:
            pass
        return redirect('resurses:myplans')

    return render(request, 'resurses/prodject-myplans.html',{'use':use, 'projd': projd,'proj':proj, 'zada':zada})

# МОЯ КОМАНДА 
@login_required(login_url='/lg/')
def myworkers(request, pk):
    cont = Cotrudniri.objects.filter(projects__id=pk).filter(tip='sotrudnic')
    contac = Cotrudniri.objects.exclude(projects__id=pk).filter(tip='sotrudnic')
    if request.method == 'POST':
        prog = Projects.objects.get(id= pk)
        if request.POST.get('form') == 'del':
            try:
                for u in request.POST.getlist('delete'):
                    cont = Cotrudniri.objects.get(index =u)
                    cont.projects.remove(prog)
                    cont.save()
                
                cont = Cotrudniri.objects.filter(projects__id=pk).filter(tip='sotrudnic')
                contac = Cotrudniri.objects.exclude(projects__id=pk).filter(tip='sotrudnic')
            except:
                pass
        elif request.POST.get('form') == 'searh':
            contac = Cotrudniri.objects.exclude(projects__id=pk).filter(tip='sotrudnic').filter(title__startswith=request.POST.get('s'))
        
        elif request.POST.get('form') == 'add':
            try:
                for u in request.POST.getlist('append'):
                    contac = Cotrudniri.objects.get(index =u)
                    contac.projects.add(prog)
                    contac.save()

                cont = Cotrudniri.objects.filter(projects__id=pk).filter(tip='sotrudnic')
                contac = Cotrudniri.objects.exclude(projects__id=pk).filter(tip='sotrudnic')
            except:
                pass
        else:
            pass
        #return redirect("/a/workers/%s/" % (pk)) 
        return render(request, 'resurses/prodject-myworcers.html',{"cont":cont , 'contac':contac })
    return render(request, 'resurses/prodject-myworcers.html',{"cont":cont , 'contac':contac })

# СОЗДАТЬ СОТРУДНИКА 
@login_required(login_url='/lg/')
def mywokrcrete(request):
    otd = Otdel.objects.all()
    compan = Compani.objects.all()
    if request.method == 'POST':
        cont = Cotrudniri()
        cont.tip = 'sotrudnic'
        cont.title  = request.POST.get('title')
        cont.doljnost = request.POST.get('doljnost')
        cont.mail = request.POST.get('mail')
        cont.ondel = otd.get(id = request.POST.get('otd'))
        cont.telnomder  = request.POST.get('telnomder')
        cont.statusof = request.POST.get('statusof')
        #if request.POST.get('compan') != '':
        #   cont.compan = Compani.objects.get(id =request.POST.get('compan')) 
    
        cont.save()
        return redirect("/a/myworcersall/" ) 
    return render(request, 'resurses/prodject-criate.html',{"otd":otd ,'compan':compan})

# ИЗМЕНИТЬ КОНТАКТ СОТРУДНИК 
@login_required(login_url='/lg/')
def mywokrcrename(request,pk):
    otd = Otdel.objects.all()
    cont = Cotrudniri.objects.get(index = pk)
    compan = Compani.objects.all()
    if request.method == 'POST':
        cont.title  = request.POST.get('title')
        cont.doljnost  = request.POST.get('doljnost')
        cont.mail  = request.POST.get('mail')
        cont.ondel = otd.get(id  = request.POST.get('otd'))
        cont.telnomder  = request.POST.get('telnomder')
        cont.statusof = request.POST.get('statusof')
        if request.POST.get('statusof') == "Уволен":
            cont.datafilds = datetime.now()
        #cont.compan = Compani.objects.get(id =request.POST.get('compan')) 
        cont.save()
        return redirect("/a/myworcersall/" ) 
    return render(request, 'resurses/prodject-criate.html',{"otd":otd ,'con':cont ,'compan':compan})

# ВСЕ КОНТАКТЫ 
@login_required(login_url='/lg/')
def myworcersall(request):
    cont = Cotrudniri.objects.filter(tip='sotrudnic')
    return render(request, 'resurses/prodject-myworcersall.html',{"cont":cont})

# ФИЛЬТР ОТДЕЛ
@login_required(login_url='/lg/')
def myotdel(request,pk):
    cont = Cotrudniri.objects.filter(ondel__id = pk).filter(tip='sotrudnic')
    return render(request, 'resurses/prodject-myworcersall.html',{"cont":cont})

#
@login_required(login_url='/lg/')
def allresorses(request):
    #if  request.user.username == 'boss':
    #    prog = Projects.objects.exclude(activedog = "Завершен")
    #else:  filter(rpproecta='Максим Пех')
    per = OtchetforMonf.objects.all()
    otd = Otdel.objects.all()
    prog = Projects.objects.exclude(activedog = "Завершен").filter(rpproecta=request.user.first_name) #request.user.first_name 
    cont = Cotrudniri.objects.exclude(statusof='Уволен').filter(tip='sotrudnic')
    return render(request, 'resurses/prodject-projects.html',{"cont":cont , 'prog':prog , 'otd':otd , 'per':per})

# РЕСУРСЫ 
@login_required(login_url='/lg/')
def resorsmatch(request):
    prog = Projects.objects.filter(rpproecta=request.user.first_name).exclude(activedog = "Завершен") #request.user.first_name
    cont = Cotrudniri.objects.exclude(statusof='Уволен').filter(tip='sotrudnic')
    if request.method == 'POST':
        ot = OtchetforMonf.objects.all().last()
        idcont = request.POST.getlist('idprog')
        d = 0 
        for i in request.POST.getlist('proz'):
            if i == '-':
                d = d+1 
            else:
                a = idcont[d].split('?')
                try:
                    pr = PriseofmyWok.objects.filter(cotrud__index=a[1]).filter(otch=ot).filter(projects__id=a[0]).get()
                    pr.title = i
                    pr.save()
                    d = d+1 
                except: 
                    a = idcont[d].split('?')
                    p = PriseofmyWok()
                    p.title = i
                    p.projects = Projects.objects.get(id=a[0])
                    p.otch = ot
                    #if Itap.objects.filter(projects_id=a[0]).exclude(act='Подписан').first() == None:
                    #    p.itap = Itap.objects.filter(projects_id=a[0]).last()
                    #else:
                    #    p.itap = Itap.objects.filter(projects_id=a[0]).exclude(act='Подписан').first() 
                    p.cotrud = Cotrudniri.objects.get(index=a[1])
                    p.save()
                    d= d+1

        return redirect("/a/allresorses/") 

    return render(request, 'resurses/prodject-projects-proz.html',{"cont":cont , 'prog':prog })

# ОТЧЕТ ЗА МЕСЯЦ + ИКСЕЛЬ - ________
@login_required(login_url='/lg/')
def rezalt(request,pk):
    #per = OtchetforMonf.objects.all(), 'per':per
    
    ot = OtchetforMonf.objects.get(id=pk)
    cont = PriseofmyWok.objects.filter(otch =ot)
    
    if request.method == 'POST':
        #wb = Workbook()
        work = os.path.join(MEDIA_ROOT ,"otchet.xlsx" )
        wb = load_workbook(filename = work)
        for sheet_name in wb.sheetnames:
            sheet = wb.get_sheet_by_name(sheet_name)
            wb.remove_sheet(sheet)
        ws = wb.create_sheet(ot.title)
        #ws.append(['Проект' ,'Предприятие','Система ','Этап','ФИО','Должность','Отдел','Процент' ])
        #wb.save("empty_book.xlsx")
        #ws = wb.active
        #ws.append([ 'Счет','Проект в РМО ',  'Организация' , 'ОрганизацияNrec' , 'Проект в Галактике','Dogovor' ,'ПроектNrec', 'ФИО' , ' ФИО Nrec ', 'Система', 'Система Nrec' , ' Этап', 'ЭтапNrec','Процент' ])
        #ws.append([ '', res.projects.title  , pd , md , res.projects.galactikanaumenovanie , res.projects.galactikaID,  res.cotrud.title ,res.cotrud.urlform , ' '.join([i.title for i in res.projects.products.all() ]) ,' ', '' , int(res.title)/100])
          

        ws.append(['Проект','dogovor' ,'Предприятие','ID','Система ','Этап','ФИО' , 'ID ','Отдел','Процент' ])
        i= 1
        for res in cont: 
            try:
                pdr = Compani.objects.get(id =res.projects.compani)
                pd = pdr.title 
                md = pdr.naimvsisteme
            except:
                pd = '-'
                md= '-'
            ws.append([res.projects.title ,res.projects.nomerdogovora  , pd , md , ' '.join([i.title for i in res.projects.products.all() ]) ,' ', res.cotrud.title ,res.cotrud.doljnost , res.cotrud.ondel.title , int(res.title)/100])
            i +=1

        save_workbook(wb, work)   
        #wb.save("media/otchet.xlsx")
        return FileResponse(open(work, 'rb'))
        #return redirect("/p/%s/" % ('otchet.xlsx'))
    return render(request, 'resurses/prodject-projects-rezalt.html',{"cont":cont, 'itm' :ot   })

# РЕСУРСЫ ПО ОТДЕЛАМ 
@login_required(login_url='/lg/')
def otdelcontacts(request,pk):
    otd = Otdel.objects.all()
    per = OtchetforMonf.objects.all()
    gor  = OtchetforMonf.objects.all().last()
    cont = PriseofmyWok.objects.filter(cotrud__ondel__id=pk)
    sot = Cotrudniri.objects.filter(ondel__id=pk).exclude(statusof='Уволен').filter(tip='sotrudnic')
    otdel = Otdel.objects.get(id=pk).title
    proj = []
    for i in sot:
        for b in  i.projects.exclude(activedog ="Завершен") :
            if b.id in proj:
                pass
            else:
                proj.append([str(b.id), str(b.title)]) 

    #komadamodel = list(PriseofmyWok.objects.filter(projects__id=pk).values_list('otch_id', flat=True).distinct())
    # .order_by('-activedog') 
    if request.method == 'POST':
        o =0
        dpds = request.POST.getlist('checid')
        for i in request.POST.getlist('proz'):
            if i == '-':
                o =o+1 
            else:
                try:
                    pr = PriseofmyWok.objects.filter(projects__id=dpds[o].split('?')[0]).filter(cotrud__index=dpds[o].split('?')[1]).filter(otch=gor).get()
                    pr.title = i
                    pr.save()
                    o =o+1 
                except:
                    pr = PriseofmyWok()
                    pr.projects = Projects.objects.get(id =dpds[o].split('?')[0])
                    pr.cotrud = Cotrudniri.objects.get(index=dpds[o].split('?')[1])
                    pr.otch = gor
                    pr.title = i
                    #if Itap.objects.filter(projects__id=dpds[o].split('?')[0]).exclude(act='Подписан').first() == None:
                    #    pr.itap = Itap.objects.filter(projects__id=dpds[o].split('?')[0]).last()
                    #else:
                    #    pr.itap = Itap.objects.filter(projects__id=dpds[o].split('?')[0]).exclude(act='Подписан').first() 
                    pr.save()
                    o =o+1 
        
        return redirect("/a/allresorses/" ) 

    return render(request, 'resurses/prodject-projects-otdel.html',{"cont":cont , 'otd':otd , 'otdel':otdel , 'sot':sot , 'proj':dict(proj) , 'per':per ,'gor':gor})

# КОНТАкты КОМПАНИИ 
@login_required(login_url='/lg/')
def companicontacts(request,pk):
    prod = Projects.objects.filter(compani=pk)
    pres = PresailProjects.objects.filter(zakazchic__id=pk) 
    comp = Compani.objects.get(id = pk)
    compne = Compani.objects.all()
    cont = Cotrudniri.objects.filter(compan=comp)
    contacts = Cotrudniri.objects.exclude(tip='cotrudcompani')
    if request.method == 'POST':

        comp.adres = request.POST.get('adres')
        comp.opisanie = request.POST.get('opisanie')
        comp.email = request.POST.get('email')
        comp.sait = request.POST.get('sait')
        comp.inn = request.POST.get('inn')
        comp.otraasal = request.POST.get('otraasal')
        comp.telefon = request.POST.get('telefon')
        comp.kontakt = request.POST.get('kontakt')
        fome = request.POST.get('holding')
        if  fome =="":
            pass
        else: 
            comp.main_comp = Compani.objects.get(id = fome)

        comp.save()
        try:
            d = 0
            for i in request.POST.getlist('fiosot'):
                if i == "":
                    d=d+1
                else:    
                    cot = Cotrudniri()
                    cot.title = i
                    cot.tip = 'cotrudcompani'
                    cot.doljnost = request.POST.getlist('doljnsotr')[d]
                    cot.compan = comp
                    cot.podraz = request.POST.getlist('podraz')[d]
                    cot.telnomder = request.POST.getlist('telsotrud')[d]
                    cot.mail = request.POST.getlist('emailsotr')[d]
                    cot.save()
                    d=d+1
        except:
            pass
        
        return redirect("/a/companicontacts/%s/" % (pk))

    return render(request, 'resurses/projects-compani-contacts.html',{ 'prod':prod, 'pres':pres, "comp":comp ,'compne':compne ,'cont':cont , 'contacts':contacts})

# КОНТАКТЫ ДРУГИХ КОМПАНИЙ
@login_required(login_url='/lg/')
def allcontactscompani(request):
    comp = Compani.objects.all()
    cont = Cotrudniri.objects.filter(tip='cotrudcompani')
    return render(request, 'resurses/projects-compani-contacts-all.html',{"comp":comp , 'cont':cont })

# ВСЕ КОМПААНИИ
@login_required(login_url='/lg/')
def allcompanis(request):
    comp = Compani.objects.all()
    return render(request, 'resurses/prodject-compani-all.html',{"cont":comp  })

# Редактирование контакта 
@login_required(login_url='/lg/')
def mcontactdetail(request,pk):
    cot = Cotrudniri.objects.get(index=pk)
    if request.method == 'POST':
        cot.title = request.POST.get('title')
        cot.doljnost = request.POST.get('doljnost')
        cot.podraz = request.POST.get('podraz')
        cot.telnomder = request.POST.get('telnomder')
        cot.mail = request.POST.get('mail')
        cot.save()
        return redirect("/a/companicontacts/%s/" % (cot.compan.id))
    return render(request, 'resurses/prodject-cont-my.html',{"cont":cot})

# РЕСУРСЫ ВСЕ ПРОЕКТЫ С КОМАНДАМИ
@login_required(login_url='/lg/')
def resorsmatc_all(request):
    prog = Projects.objects.exclude(activedog = "Завершен") #request.user.first_name
    cont = Cotrudniri.objects.exclude(statusof='Уволен').filter(tip='sotrudnic')
    if request.method == 'POST':
        ot = OtchetforMonf.objects.all().last()
        
        idcont = request.POST.getlist('idprog')
        d = 0 
        for i in request.POST.getlist('proz'):
            if i == '-':
                d = d+1 
            else:
                a = idcont[d].split('?')
                try:
                    pr = PriseofmyWok.objects.filter(cotrud__index=a[1]).filter(otch=ot).filter(projects__id=a[0]).get()
                    pr.title = i
                    pr.save()
                    d = d+1 
                except: 
                    a = idcont[d].split('?')
                    p = PriseofmyWok()
                    p.title = i
                    p.projects = Projects.objects.get(id=a[0])
                    p.otch = ot 
                    #if Itap.objects.filter(projects_id=a[0]).exclude(act='Подписан').first() == None:
                    #    p.itap = Itap.objects.filter(projects_id=a[0]).last()
                    #else:
                    #    p.itap = Itap.objects.filter(projects_id=a[0]).exclude(act='Подписан').first() 
                    p.cotrud = Cotrudniri.objects.get(index=a[1])
                    p.save()
                    d= d+1

        return redirect("/a/allresorses/") 

    return render(request, 'resurses/prodject-projects-proz.html',{"cont":cont , 'prog':prog })


# Матрица общаяя
@login_required(login_url='/lg/')
def otdelcontacts_all(request,pk):
    gor  = OtchetforMonf.objects.get(id = pk)
    sot = Cotrudniri.objects.exclude(statusof='Уволен').filter(tip='sotrudnic')
    proj = Projects.objects.all()
    return render(request, 'resurses/prodject-projects-otdel-all.html',{ 'sot':sot , 'proj':proj  ,'gor':gor})


# СОГЛОСОВАНИЕ РЕСУРСОВ 
@login_required(login_url='/lg/')
def soglosovanie_resursu(request):
    # request.user.first_name Максим Пех
    prog = Projects.objects.filter(rpproecta=request.user.first_name).exclude(activedog = "Завершен").order_by('-activedog') #request.user.first_name
    cont = Cotrudniri.objects.exclude(statusof='Уволен').filter(tip='sotrudnic').filter(projects__in=prog).distinct()
    gor  = OtchetforMonf.objects.all().last()
    if request.method == 'POST':
        o =0
        dpds = request.POST.getlist('checid')
        for i in request.POST.getlist('proz'):
            if i == '-':
                o =o+1 
            else:
                try:
                    pr = PriseofmyWok.objects.filter(projects__id=dpds[o].split('?')[0]).filter(cotrud__index=dpds[o].split('?')[1]).filter(otch=gor).get()
                    pr.title = i
                    pr.soglos = 'soglos'
                    pr.save()
                    o =o+1 
                except:
                    pr = PriseofmyWok()
                    pr.projects = Projects.objects.get(id =dpds[o].split('?')[0])
                    pr.cotrud = Cotrudniri.objects.get(index=dpds[o].split('?')[1])
                    pr.otch = gor
                    pr.soglos = 'soglos'
                    pr.title = i
                    #if Itap.objects.filter(projects__id=dpds[o].split('?')[0]).exclude(act='Подписан').first() == None:
                    #    pr.itap = Itap.objects.filter(projects__id=dpds[o].split('?')[0]).last()
                    #else:
                    #    pr.itap = Itap.objects.filter(projects__id=dpds[o].split('?')[0]).exclude(act='Подписан').first() 
                    pr.save()
                    o =o+1 
        
        return redirect("/a/allresorses/" ) 

 
    return render(request, 'resurses/prodject-projects-soglos.html',{'prog':prog ,'cont' :cont})






@login_required(login_url='/lg/')
def makemigratinonresors(request, pk):
    gor  = OtchetforMonf.objects.all().last()

    for i in  PriseofmyWok.objects.exclude(projects__activedog = "Завершен").filter(otch_id= pk):
        pr = PriseofmyWok()
        pr.projects = i.projects
        pr.cotrud = i.cotrud
        pr.otch = gor
        pr.title = i.title
        #if Itap.objects.filter(projects__id=dpds[o].split('?')[0]).exclude(act='Подписан').first() == None:
        #    pr.itap = Itap.objects.filter(projects__id=dpds[o].split('?')[0]).last()
        #else:
        #    pr.itap = Itap.objects.filter(projects__id=dpds[o].split('?')[0]).exclude(act='Подписан').first() 
        pr.save() 
    return redirect("/a/allresorses/") 





#ИКСЕЛЬ - ________
@login_required(login_url='/lg/')
def prog(request):
    #per = OtchetforMonf.objects.all(), 'per':per
    prog = Projects.objects.exclude(activedog = "Завершен").exclude(activedog = "Системный")
    
    #wb = Workbook()
    work = os.path.join(MEDIA_ROOT ,"otchet.xlsx" )
    wb = load_workbook(filename = work)
    for sheet_name in wb.sheetnames:
        sheet = wb.get_sheet_by_name(sheet_name)
        wb.remove_sheet(sheet)
    ws = wb.create_sheet('Лист1')
    ws.append(['название ' , 'организация', 'РП' , 'номер договора' ])
   
    i= 1
    for res in prog: 
        try:
            pdr = Compani.objects.get(id =res.projects.compani).title
        except:
            pdr = ''

        ws.append([res.title , pdr ,res.rpproecta  , res.nomerdogovora])
        i +=1

    save_workbook(wb, work)   
    #wb.save("media/otchet.xlsx")
    return FileResponse(open(work, 'rb'))
 









