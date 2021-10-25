from django.shortcuts import render , redirect
from django.contrib.auth.decorators import login_required
# Create your views here.
from resurses.models import  Cotrudniri , Otdel
from ushetvremeni.models import UshetResursov

from openpyxl.styles import Font 
from openpyxl.styles.borders import Border, Side
from openpyxl import Workbook
from django.http import FileResponse
from django.conf.global_settings import MEDIA_ROOT

import datetime
import calendar

import os

@login_required(login_url='/lg/')
def my_otdel_plan(request, pk):
    year = datetime.datetime.now().year
    month = datetime.datetime.now().month
    otd = Otdel.objects.all()
    cont = Cotrudniri.objects.exclude(statusof='Уволен').filter(tip='sotrudnic').filter(ondel__id=pk)
    data = calendar.Calendar(firstweekday=0).monthdays2calendar(year, month) 
    if request.method == 'POST':
        work = os.path.join(MEDIA_ROOT ,"yshetres.xlsx" ) 
        #work = 'yshetres.xlsx'
        wb = Workbook()
        wb.create_sheet(title = 'Учет', index = 0)
        sheet = wb['Учет']
        font = Font(size=12, bold=True)
        font2 = Font(size=8, italic=True)
        font3 = Font(bold=True)

        sheet.append([ '' ])
        sheet.append([ '' , '' ,'ТАБЕЛЬ УЧЕТА РАБОЧЕГО ВРЕМЕНИ СОТРУДНИКОВ ПОДРАЗДЕЛЕНИЯ' ])
        sheet.append([ '' ])     
        sheet.append(['' , '', Otdel.objects.get(id=pk).title ])
        sheet.append(['' , '', '                      (название подразделения) ' ])
        sheet.append([ '' ])   
        sheet.append([ '' ,'',f"за период с  { request.POST.getlist('datavolues')[0] }  по  { request.POST.getlist('datavolues')[1] } "])
        sheet.append([ '' ])        
        sheet.append([ '','№', 'Ф.И.О.', '  ПН' ,'  ВТ','  СР', '  ЧТ', "  ПТ" , "  СБ",  '  ВС'])  
        wek = request.POST.get('wek')
        proz = request.POST.getlist('proz')
        n = 0 
        k = 1
        for i in request.POST.getlist('sotr'):
            obj = UshetResursov.objects.get(title=wek, year=year, cotrud__index=i )
            obj.rezalt = ','.join([ i for i in proz[n:n+7] ])
            obj.save()
            sheet.append(list([ '', k  , obj.cotrud.title ]  + [ i for i in proz[n:n+7] ]))  
            n += 7 
            k += 1

        sheet.append([ '' ]) 
        sheet.append([ '' ]) 
        sheet.append([ '' ,f'Подпись руководителя подразделения________________________  ___________________  { datetime.datetime.today().strftime("%d.%m.%Y") }'])     
        sheet.append([ '',  '', '', '(подпись)', '', '(расшифровка)' ]) 
        
        sheet.column_dimensions['B'].width = 3
        sheet.column_dimensions['C'].width = 40
        sheet.column_dimensions['D'].width = 6
        sheet.column_dimensions['E'].width = 6
        sheet.column_dimensions['F'].width = 6
        sheet.column_dimensions['G'].width = 6
        sheet.column_dimensions['H'].width = 6
        sheet.column_dimensions['I'].width = 6
        sheet.column_dimensions['J'].width = 6

        thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin' ))

        for x in range(9 , 9+ k ):
            for y in range(2 , 11):
                sheet.cell(row=x, column=y).border = thin_border
        
        sheet['C2'].font = font
        sheet['C5'].font = font2
        sheet['C7'].font = font3
        sheet[f'B{11 + k}'].font = font3
        sheet[f'D{12 + k}'].font = font2
        sheet[f'F{12 + k}'].font = font2

        wb.save(work)
        return FileResponse(open(work, 'rb'))

        #return redirect(f"/ut/{pk}" ) 
    
    return render(request, 'ushetvremeni/tabelvremeni.html',{ 'cont':cont , 'data': data  , 'year':year ,'month':month  , 'otd': otd})










