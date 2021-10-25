from django.shortcuts import render , redirect
from django.contrib.auth.decorators import login_required
#import smtplib
from django.core.mail import send_mail , send_mass_mail
from projects.models import Projects , Itap , Komanda , Zadania , Dopsog , Zadachirp, Mani
from resurses.models import Otdel , Cotrudniri , Compani
from preseil.models import PresailProjects , StadiiProdaj
from django.contrib.auth.models import User
from django.http import HttpResponseRedirect
from .forms import UploadDocumentForm , UploadFileForm
import openpyxl

# Create your views here.


@login_required(login_url='/lg/')
def operplan(request, pk):
    projects = Projects.objects.get(id = pk)
    itap= Itap.objects.filter(projects__title=projects)

    form = UploadFileForm()    
    if request.method == 'POST':
        form = UploadFileForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                wb = openpyxl.load_workbook(filename = request.FILES['file'],data_only = True)
                svv = wb['Лист1']
                i =3 
                for itm in itap:
                    """
                    mo = Mani()
                    mo.title = itm.title
                    mo.itap = itm
                    mo.save()
                    """
                    while str(svv['A' + str(i)].value) !='None' or  str(svv['B' + str(i)].value) !='None':
                        if  len(str(svv['A' + str(i)].value)) == 1:
                            i=i+1
                            break
                        else: 
                            zad = Zadania()
                            zad.itap = itm
                            zad.level = str(svv['A' + str(i)].value)
                            zad.title = str(svv['B' + str(i)].value)
                            dat =str(svv['C' + str(i)].value).split(' ')
                            dat1 =str(svv['D' + str(i)].value).split(' ')
                            zad.data = dat[0]
                            zad.dataend = dat1[0]
                            zad.rezult = str(svv['F' + str(i)].value)
                            zad.prozes = '0%'
                            zad.save()
                            i=i+1 
            except:
                pass
        

        return redirect('glavarp:admins')


    return render(request, 'glavarp/prodject-op.html', {'form': form})

@login_required(login_url='/lg/')
def main_page(request):
    use = User.objects.all()
    if request.method == 'POST':
        qw = request.POST.getlist('noms')
        return redirect('projects:create',pk = qw[0])#projects:create
    return render(request, 'glavarp/list.html',{'use':use})

@login_required(login_url='/lg/')
def post_page(request):
    use = User.objects.all()
    #mails = ('makeev.21@mail.ru','makeev.21@mail.ru')
    mess = []
    for e in use:
        if e.email == '':
            print('1')
        else:
            ot = ('Отчет по проектам','Добрый день,коллеги. Прошу обновить статус отчет по Вашим проектам на портале:','galaktika.ot@gmail.com', [ e.email ])
            mess.append(ot)
    """
    send_mail('Отчет по проектам','Добрый день,коллеги.  \r\n  \r\n Прошу обновить статус отчет по Вашим проектам на портале:','galaktika.ot@gmail.com',
        ['makeev.21@mail.ru'],
        fail_silently=False,)
        """
    send_mass_mail((mess), fail_silently=False)

    return render(request, 'glavarp/list.html', {'post':" Сообщения отправленны "})

@login_required(login_url='/lg/')
def projectstest(request):
    proj = Projects.objects.all()
    return render(request, 'glavarp/index.html',{"proj":proj})

@login_required(login_url='/lg/')
def docwith(request):
    use = User.objects.all()

    if request.method == 'POST':
        p = Projects()
        p.rpproecta = request.POST.get('location')
        p.title =  request.POST.get('nazvaniesis')
        p.nomerdogovora = request.POST.get('nomerdog')
        p.tipdogovora = request.POST.get('tipproect')
        p.save()
        return redirect('glavarp:admins')
    return render(request, 'glavarp/prodject-criate.html', {'use': use})

@login_required(login_url='/lg/')
def docwith_old(request):
    if request.method == 'POST':
        form = UploadFileForm(request.POST, request.FILES)
        if form.is_valid():
            
            wb = openpyxl.load_workbook(filename = request.FILES['file'],data_only = True)
            svv = wb['Отчет ГД'] #название книги 

            i=4 
            while str(svv['D' + str(i)].value) !='None' or  str(svv['F' + str(i)].value) !='None':
                if str(svv['A' + str(i)].value) !='None':
                   # print(str(svv['A' + str(i)].value), str(svv['C' + str(i)].value))
                    
                    ff=str(svv['A' + str(i)].value).split('|')
                    proj = Projects()
                    proj.title = ff[0]
                    proj.rpproecta = ff[1]
                    proj.tipdogovora = str(svv['B' + str(i)].value)
                    proj.nomerdogovora = ff[2]
                    proj.save()
                    proje = Projects.objects.get(title=ff[0])

                    itap = Itap()
                    itap.title = str(svv['C' + str(i)].value)
                    itap.itapnaim = str(svv['D' + str(i)].value)
                   # dat =str(svv['E' + str(i)].value).split(' ')
                    
                    try:
                        if str(svv['F' + str(i)].value) != "None":
                            dat =str(svv['F' + str(k)].value).split(' ')
                            dat1 =str(svv['E' + str(k)].value).split(' ')
                            itap.datastart = dat[0]
                            itap.datastop = dat1[0]
                        else:
                            pass
                    except:
                        pass

                    itap.projects = proje
                    itap.zadolkjnost = ''
                    try:
                        x = str(svv['K' + str(i)].value).split('Ирина Степанова:')
                        itap.komentstep = x[1]
                        itap.komentrp = x[0]
                    except:
                        itap.komentstep = ''
                        itap.komentrp = str(svv['K' + str(i)].value)
                    
                    itap.save()

                    k = i 
                    k = k+ 1
                    if str(svv['A' + str(k)].value) =='None': # 2
                        #print(str(svv['A' + str(i)].value), str(svv['C' + str(k)].value))
                        itap = Itap()
                        itap.title = str(svv['C' + str(k)].value)
                        itap.itapnaim = str(svv['D' + str(k)].value)
                        #dat =str(svv['E' + str(k)].value).split(' ')
                        try:
                            if str(svv['F' + str(k)].value) != "None":
                                dat =str(svv['F' + str(k)].value).split(' ')
                                dat1 =str(svv['E' + str(k)].value).split(' ')
                                itap.datastart = dat[0]
                                itap.datastop = dat1[0]
                            else:
                                pass
                        except:
                            pass
                        itap.projects = proje
                        itap.zadolkjnost = ''
                        try:
                            x = str(svv['K' + str(k)].value).split('Ирина Степанова:')
                            itap.komentstep = x[1]
                            itap.komentrp = x[0]
                        except:
                            itap.komentstep = ''
                            itap.komentrp = str(svv['K' + str(k)].value)
                        
                        itap.save()
                        k = k+ 1
                        if str(svv['A' + str(k)].value) =='None': # 3
                            print(str(svv['A' + str(i)].value), str(svv['C' + str(k)].value))
                            itap = Itap()
                            itap.title = str(svv['C' + str(k)].value)
                            itap.itapnaim = str(svv['D' + str(k)].value)
                            dat =str(svv['F' + str(k)].value).split(' ')
                            dat1 =str(svv['E' + str(k)].value).split(' ')
                            itap.datastart = dat[0]
                            itap.datastop = dat1[0]
                            itap.projects = proje
                            itap.zadolkjnost = ''
                            try:
                                x = str(svv['K' + str(k)].value).split('Ирина Степанова:')
                                itap.komentstep = x[1]
                                itap.komentrp = x[0]
                            except:
                                itap.komentstep = ''
                                itap.komentrp = str(svv['K' + str(k)].value)
                            
                            itap.save()


                            k = k+ 1
                            if str(svv['A' + str(k)].value) =='None': # 4
                                print(str(svv['A' + str(i)].value), str(svv['C' + str(k)].value))

                                itap = Itap()
                                itap.title = str(svv['C' + str(k)].value)
                                itap.itapnaim = str(svv['D' + str(k)].value)
                                dat =str(svv['F' + str(k)].value).split(' ')
                                dat1 =str(svv['E' + str(k)].value).split(' ')
                                itap.datastart = dat[0]
                                itap.datastop = dat1[0]
                                itap.projects = proje
                                itap.zadolkjnost = ''
                                try:
                                    x = str(svv['K' + str(k)].value).split('Ирина Степанова:')
                                    itap.komentstep = x[1]
                                    itap.komentrp = x[0]
                                except:
                                    itap.komentstep = ''
                                    itap.komentrp = str(svv['K' + str(k)].value)
                                
                                itap.save()

                                k = k+ 1
                                if str(svv['A' + str(k)].value) =='None': # 5
                                    print(str(svv['A' + str(i)].value), str(svv['C' + str(k)].value))

                                    itap = Itap()
                                    itap.title = str(svv['C' + str(k)].value)
                                    itap.itapnaim = str(svv['D' + str(k)].value)
                                    dat =str(svv['F' + str(k)].value).split(' ')
                                    dat1 =str(svv['E' + str(k)].value).split(' ')
                                    itap.datastart = dat[0]
                                    itap.datastop = dat1[0]
                                    itap.projects = proje
                                    itap.zadolkjnost = ''
                                    try:
                                        x = str(svv['K' + str(k)].value).split('Ирина Степанова:')
                                        itap.komentstep = x[1]
                                        itap.komentrp = x[0]
                                    except:
                                        itap.komentstep = ''
                                        itap.komentrp = str(svv['K' + str(k)].value)
                                    
                                    itap.save()


                                    k = k+ 1
                                    if str(svv['A' + str(k)].value) =='None': # 6
                                        print(str(svv['A' + str(i)].value), str(svv['C' + str(k)].value))

                                        itap = Itap()
                                        itap.title = str(svv['C' + str(k)].value)
                                        itap.itapnaim = str(svv['D' + str(k)].value)
                                        dat =str(svv['F' + str(k)].value).split(' ')
                                        
                                        itap.datastart = dat[0]
                                        itap.datastop = dat[0]
                                        itap.projects = proje
                                        itap.zadolkjnost = ''
                                        try:
                                            x = str(svv['K' + str(k)].value).split('Ирина Степанова:')
                                            itap.komentstep = x[1]
                                            itap.komentrp = x[0]
                                        except:
                                            itap.komentstep = ''
                                            itap.komentrp = str(svv['K' + str(k)].value)
                                        
                                        itap.save()

                                        k = k+ 1
                                        if str(svv['A' + str(k)].value) =='None': # 7
                                            print(str(svv['A' + str(i)].value), str(svv['C' + str(k)].value))
                                        
                                            itap = Itap()
                                            itap.title = str(svv['C' + str(k)].value)
                                            itap.itapnaim = str(svv['D' + str(k)].value)
                                            dat =str(svv['F' + str(k)].value).split(' ')
                                            dat1 =str(svv['E' + str(k)].value).split(' ')
                                            itap.datastart = dat[0]
                                            itap.datastop = dat1[0]
                                            itap.projects = proje
                                            itap.zadolkjnost = ''
                                            try:
                                                x = str(svv['K' + str(k)].value).split('Ирина Степанова:')
                                                itap.komentstep = x[1]
                                                itap.komentrp = x[0]
                                            except:
                                                itap.komentstep = ''
                                                itap.komentrp = str(svv['K' + str(k)].value)
                                            
                                            itap.save()

                        
                                        else:
                                            pass
                                    else:
                                        pass
                                else:
                                    pass
                            else:
                                pass
                        else:
                            pass
                    else:
                        pass

                    i = i+1
                elif str(svv['A' + str(i)].value) == 'None':
                    i = i+1 
                    
                else: 
                    print(' mistake')
                    i = i+1 
                    pass
            return redirect('glavarp:admins')
    else:
        form = UploadFileForm()        

    return render(request, 'glavarp/prodject-criate.html', {'form': form})

@login_required(login_url='/lg/')
def adminmain(request):
    prod = Projects.objects.all()
    pr = len(prod)
    return render(request,'glavarp/prodject-main.html',{"pr":pr})

@login_required(login_url='/lg/')
def operplans(request):
    

    return render(request,'glavarp/prodject-main.html')

@login_required(login_url='/lg/')   
def migratio_to_br(request):


    return render(request,'glavarp/prodject-main.html')

@login_required(login_url='/lg/')
def cotrudniki(request):
    if request.method == 'POST':
        form = UploadFileForm(request.POST, request.FILES)
        if form.is_valid():
            
            wb = openpyxl.load_workbook(filename = request.FILES['file'],data_only = True)
            svv = wb['Лист1'] #название книги 
            i=4 
            while svv['A' + str(i)].value != None or  svv['C' + str(i)].value != None:
                if Cotrudniri.objects.filter(title=str(svv['A' + str(i)].value)).exists() != True :
                    try:
                        ot = Otdel.objects.get(title=str(svv['E' + str(i)].value)) 
                        c = Cotrudniri()
                        c.title = str(svv['A' + str(i)].value)
                        c.urlform = str(svv['B' + str(i)].value)
                        c.ondel = ot
                        c.tip = "sotrudnic"
                        c.save()
                    except:
                        ot = Otdel()
                        ot.title = str(svv['E' + str(i)].value)
                        #ot.rpotdela= str(svv['D' + str(i)].value)
                        ot.save()
                        c = Cotrudniri()
                        c.title = str(svv['A' + str(i)].value)
                        c.urlform = str(svv['B' + str(i)].value)
                        c.ondel = ot
                        c.tip = "sotrudnic"
                        c.save() 
                else:
                    #print(svv['A' + str(i)].value)
                    pass
                i=i+1
    else:
        form = UploadFileForm()  
    return render(request, 'glavarp/prodject-contacts.html', {'form': form})

@login_required(login_url='/lg/')
def mcontactdetail(request):


    return render(request, 'glavarp/prodject-cont-my.html')

@login_required(login_url='/lg/')
def compani(request):
    if request.method == 'POST':
        form = UploadFileForm(request.POST, request.FILES)
        if form.is_valid():
            
            wb = openpyxl.load_workbook(filename = request.FILES['file'],data_only = True)
            svv = wb['Лист1'] #название книги 
            i=3 
            while svv['B' + str(i)].value != None or  svv['C' + str(i)].value != None:
                comp = Compani()
                comp.title = svv['C' + str(i)].value 
                comp.adres = svv['J' + str(i)].value 
                comp.opisanie = svv['D' + str(i)].value 
                comp.email = svv['L' + str(i)].value 
                comp.sait = svv['M' + str(i)].value 
                comp.inn = svv['I' + str(i)].value 
                comp.otraasal = svv['G' + str(i)].value 
                comp.telefon = svv['K' + str(i)].value 
                comp.kontakt = svv['E' + str(i)].value 
                comp.save()
                i=i+1
    else:
        form = UploadFileForm()  
    return render(request, 'glavarp/prodject-contacts.html', {'form': form})

@login_required(login_url='/lg/')
def presaiall(request):
    if request.method == 'POST':
        form = UploadFileForm(request.POST, request.FILES)
        if form.is_valid():
            
            wb = openpyxl.load_workbook(filename = request.FILES['file'],data_only = True)
            svv = wb['Лист1'] #название книги 
            i=3 
            while svv['B' + str(i)].value != None or  svv['C' + str(i)].value != None:
                press = PresailProjects()
                press.title = svv['C' + str(i)].value 
                press.polzvt = svv['F' + str(i)].value 
                press.status = 'Активный'
                try:
                    press.zakazchic = Compani.objects.get(title=svv['C' + str(i)].value )
                except :
                    pass
                   
                try:
                    press.menedjer = User.objects.get(first_name=svv['E' + str(i)].value )
                except:
                    pass
                press.save()

                st = StadiiProdaj()
                st.title = svv['C' + str(i)].value 
                st.priail = press
                st.save()
                i=i+1
    else:
        form = UploadFileForm()  
    return render(request, 'glavarp/prodject-contacts.html', {'form': form})






